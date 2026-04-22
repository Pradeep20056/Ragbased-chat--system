"""
evaluate_full.py
────────────────
Runs 6 structured AI evaluations for a bidder and saves:
  - 6 individual JSON files
  - 1 consolidated Excel workbook
  ...all under  ./<bidder_name>evaluation/

Usage:
    python evaluate_full.py --bidder "Bidder2 offer.zip"
"""

import argparse
import json
import logging
import os
import re
from pathlib import Path

from dotenv import load_dotenv
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from langchain_ollama import OllamaEmbeddings, ChatOllama
from langchain_postgres import PGVector
from langchain_core.prompts import PromptTemplate

from models import (
    ContractsTechEvaluation,
    ContractsCommercialEvaluation,
    MaterialsPQExperience,
    MaterialsPQFinancial,
    MaterialsTechnicalEvaluation,
    MaterialsCommercialEvaluation,
)

# ── Bootstrap ────────────────────────────────────────────────
load_dotenv()
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

CONNECTION_STRING = "postgresql+psycopg://admin:password@localhost:5444/tender_eval"
COLLECTION_NAME   = "bidder_documents"

# ── Evaluation section definitions ───────────────────────────
EVALUATIONS = [
    {
        "id":     "a",
        "label":  "(a) PQC Exp. & Technical Evaluation (Contracts)",
        "schema": ContractsTechEvaluation,
        "search_query": "work order experience criteria technical qualification completion certificate executed value industry petroleum petrochemical",
        "instruction": """
You must fill EVERY field below from the bidder's submitted documents.

FIELDS TO EXTRACT:
1.1  similar_nature_of_job           → Nature of similar jobs the bidder has done (e.g., electrical, instrumentation).
1.2  value_of_work_order_required    → WO value submitted (1x/2x/3x format, in Rs. Lakhs).
1.3  details_of_work_order_submitted → Full details of the WO submitted.
1.4  work_order_number_and_date      → WO number and date.
1.5  nature_of_industry              → Industry type of WO issuer (Petroleum / Petrochemical / Refinery / Power).
1.6  name_of_issuer                  → Name of the company that issued the WO.
1.7  completion_certificate_details  → Details of completion certificate (cert number, date, from whom).
1.8  executed_value_as_per_completion_certificate → Executed value in Rs. Lakhs from completion certificate.
1.9  value_considered_for_technical_evaluation    → Value of order considered for TE (in Rs. Lakhs).
1.10 annualization_of_value          → Annualized value for ARC jobs, or "Not Applicable".
1.11 subcontract_approval_submitted  → Subcontract approval / end-user certificate: "Yes", "No", "Not Applicable".
1.12 work_order_meeting_experience_criteria → The specific WO that meets experience criteria.
2    additional_technical_pqc        → Any extra PQC requirements met or "None".
3    deviations                      → List of deviations from tender requirements (empty [] if nil).
4    query_to_be_raised              → List of queries/clarifications needed from the bidder.
5    technical_acceptance_status     → "Accepted", "Under Query", or "Rejected".
6    reason_for_rejection            → Reason if rejected, else "Not Applicable".

RULE: If a value is not found in the context, write "Not Found in Documents" — do NOT leave blank.
""",
    },
    {
        "id":     "b",
        "label":  "(b) PQC Fin. & Commercial Evaluation (Contracts)",
        "schema": ContractsCommercialEvaluation,
        "search_query": "annual turnover financial statements networth EPF ESI EMD formats submission MSE MII blacklisting commercial evaluation",
        "instruction": """
You must fill EVERY field below from the bidder's submitted documents.

FIELDS TO EXTRACT:
1   technical_qualification       → "Qualified", "Not Qualified", or "Under TE".
2   integrity_pact                → "Submitted (Page: X)", "Not Submitted", or "Not Applicable".
3   emd_applicable                → EMD amount and details, or "EMD exemption as MSE".
3a  emd_in_form_of_bg             → Bank Guarantee validity, or "Not Applicable".
4   emd_sent_to_finance           → "Yes" or "Not Applicable".
5a  annual_turnover_2021_22       → Turnover FY2021-22 in Rs. Lakhs.
5b  annual_turnover_2022_23       → Turnover FY2022-23 in Rs. Lakhs.
5c  annual_turnover_2023_24       → Turnover FY2023-24 in Rs. Lakhs.
6   share_capital                 → Share Capital in Rs. Lakhs.
6   reserve_and_surplus           → Reserve & Surplus in Rs. Lakhs.
6   loss                          → Loss in Rs. Lakhs (write "Nil" if no loss).
6   networth                      → Net Worth = Capital + Reserves - Loss. State "Positive" or "Negative".
7   epf_code_number               → EPF Code Number.
8   esi_code_number               → ESI Code Number.
9   cpcl_vendor_code              → CPCL Vendor Code or "Not Available".
10  power_of_attorney_submitted   → "Submitted" or "Not Submitted".
11a format_a_submitted            → "Submitted" or "Not Submitted".
11b format_b_submitted            → "Submitted" or "Not Submitted".
11c format_c_submitted            → "Submitted" or "Not Submitted".
11d format_d_submitted            → "Submitted" or "Not Submitted".
11e format_e_submitted            → "Submitted" or "Not Submitted".
11f format_f_submitted            → "Submitted" or "Not Submitted".
11g format_g_submitted            → "Submitted" or "Not Submitted".
11h format_h_submitted            → "Submitted" or "Not Submitted".
11i format_i_submitted            → "Submitted" or "Not Submitted".
11j format_j_submitted            → "Submitted", "Not Submitted", or "Not Applicable (CPCL VC available)".
11k format_k_submitted            → "Submitted" or "Not Submitted".
11l appendix_iia_submitted        → "Submitted" or "Not Submitted".
12  mse_status                    → MSE details: category, type, reservation.
13  mii_status                    → "Category 1", "Category 2", "Category 3", or "Not Applicable".
14  blacklisting_sap_cppp_gem     → Blacklisting in SAP/CPPP/GeM portals.
15  blacklisting_in_gst_portal    → "No" or blacklisting details.
16  deviations                    → List of commercial deviations (empty [] if Nil).
17  corrigendums_signed_submitted → "Submitted", "Not Submitted", or "Not Applicable".
18  queries_to_be_raised          → List of queries to raise.
19  commercial_evaluation_status  → "Qualified", "Not Qualified", or "Under Query".
20  reason_for_rejection          → Reason if rejected, else "Not Applicable".

RULE: If a value is not found in the context, write "Not Found in Documents" — do NOT leave blank.
""",
    },
    {
        "id":     "c",
        "label":  "(c) Materials - PQ Experience Criteria",
        "schema": MaterialsPQExperience,
        "search_query": "purchase order PO supply experience criteria GST invoice delivery challan commissioning proof of supply materials",
        "instruction": """
You must fill EVERY field below from the bidder's submitted documents.

FIELDS TO EXTRACT:
1a  po_item_description           → Description of the similar item in the Purchase Order.
1b  po_number                     → Purchase Order number.
1c  po_acceptable_date            → PO date (must be after the tender specified date).
1d  po_value                      → PO value in Rs. Lakhs.
1e  po_issuer_name                → Name of the organization that issued the PO.
1f  po_receiver_name              → Name of the bidder who received the PO.
1g  po_issuer_type_of_industry    → Industry type of PO issuer (Petroleum/Petrochemical/Power/Refinery).
1h  po_supplied_value             → Value of supplies made (1x/2x/3x PO tier with Rs. Lakhs).
1i  proof_of_supply               → Proof submitted: GST Invoices, Delivery Challans, or Completion Certificates.
1j  supplied_within_india         → "Applicable" or "Not Applicable".
1k  commissioned                  → "Applicable" or "Not Applicable".
2   additional_pqc                → Additional PQC details or "Not Applicable".
    queries_to_be_raised          → List of queries to raise.
    pqc_experience_status         → "Qualified", "Not Qualified", or "Query to be raised".
    reason_for_rejection          → Reason if not qualified, else "Not Applicable".

RULE: If a value is not found in the context, write "Not Found in Documents" — do NOT leave blank.
""",
    },
    {
        "id":     "d",
        "label":  "(d) Materials - PQ Financial Criteria",
        "schema": MaterialsPQFinancial,
        "search_query": "annual turnover balance sheet networth financial year 2021 2022 2023 2024 profit loss reserves capital",
        "instruction": """
You must fill EVERY field below from the bidder's submitted financial documents.

FIELDS TO EXTRACT:
1   annual_turnover_applicable    → Whether turnover requirement is applicable and required amount in Rs. Lakhs.
1a  annual_turnover_2021_22       → Turnover FY2021-22 in Rs. Lakhs (from ITR/audited financials).
1b  annual_turnover_2022_23       → Turnover FY2022-23 in Rs. Lakhs.
1c  annual_turnover_2023_24       → Turnover FY2023-24 in Rs. Lakhs.
2   positive_networth_for_latest_fy → "Applicable - Positive", "Applicable - Negative", or "Not Applicable".
    networth_value                → Net Worth value in Rs. Lakhs for the latest FY.
    queries_to_be_raised          → List of queries to raise.
    pq_financial_status           → "Qualified", "Not Qualified", or "Query to be raised".
    reason_for_rejection          → Reason if not qualified, else "Not Applicable".

RULE: If a value is not found in the context, write "Not Found in Documents" — do NOT leave blank.
""",
    },
    {
        "id":     "e",
        "label":  "(e) Materials - Technical Evaluation",
        "schema": MaterialsTechnicalEvaluation,
        "search_query": "technical specification signed sealed NIL deviation statement technical compliance user department requirement",
        "instruction": """
You must fill EVERY field below from the bidder's submitted documents.

FIELDS TO EXTRACT:
1   technical_specification_signed_and_sealed   → "Yes - Signed & Sealed", "No", or "Not Found".
2   nil_deviation_statement_signed_and_sealed   → "Yes - Signed & Sealed", "No", or "Not Found".
3   additional_user_department_requirement      → Additional requirements from user/indenter, or "None".
4   deviations                                  → List of technical deviations (empty [] if nil).
    query_to_be_raised                          → List of technical queries to raise.
    technical_evaluation_status                 → "Qualified", "Not Qualified", or "Query to be raised".
    reason_for_rejection                        → Reason if not qualified, else "Not Applicable".

RULE: If a value is not found in the context, write "Not Found in Documents" — do NOT leave blank.
""",
    },
    {
        "id":     "f",
        "label":  "(f) Materials - Commercial Evaluation (CBA)",
        "schema": MaterialsCommercialEvaluation,
        "search_query": "vendor code contact person email mobile EMD MSE UDYAM MII local content GST blacklisting integrity pact deviations validity declarations",
        "instruction": """
You must fill EVERY field below from the bidder's submitted documents.

FIELDS TO EXTRACT:
1   vendor_code                        → Vendor Code of the bidder.
2   contact_person                     → Name of the contact person.
3   mobile_number                      → Mobile number of the contact person.
4   email_id                           → Email ID of the contact person.
5   emd_applicability                  → "Applicable" or "Not Applicable".
6   emd_details                        → EMD amount, form (BG/DD/etc.), reference number.
7   emd_exemption_reason               → Reason for exemption (MSE/NSIC) or "Not Applicable".
8   mse_preference_applicability       → "Applicable" or "Not Applicable".
9   mse_preference_applied_in_gem      → "Applied" or "Not Applied".
10  udyam_number                       → UDYAM number or "Not Applicable".
11  mse_category                       → "Micro", "Small", "Medium", or "Not Applicable".
12  mse_pp_verification                → "Verified", "To be Verified", or "Not Applicable".
13  mse_purchase_preference_eligibility → "Eligible" or "Not Eligible".
14  mii_preference_applicability       → "Applicable" or "Not Applicable".
15  mii_preference_applied_in_gem      → "Applied" or "Not Applied".
16  mii_local_content_declaration      → "Submitted", "To be Verified", or "Not Applicable".
17  local_content_percentage           → % of local content declared.
18  mii_purchase_preference_eligibility → "Eligible" or "Not Eligible".
19  integrity_pact                     → "Applicable - Submitted", "Applicable - Not Submitted", "Not Applicable".
20  confidentiality_clause_declaration → "Signed & Sealed" or "Not Submitted".
21  holiday_listing_declaration        → "Signed & Sealed" or "Not Submitted".
22  land_border_sharing_declaration    → "Signed & Sealed" or "Not Submitted".
23  nil_deviations_declaration         → "Signed & Sealed" or "Not Submitted".
24  details_of_deviations              → List of commercial deviations (empty [] if nil).
25  deviations_acceptance              → "Accepted", "Not Accepted", or "Not Applicable".
26  validity                           → Bid validity (should be 120 days from final bid due date).
27  gst_number                         → GST Registration Number.
28  gst_filing_status                  → "Regular", "Irregular", or "Not Found".
29  blacklisted_in_cpcl_mopng         → "No" or blacklisting details.
    query_to_be_raised                 → List of commercial queries to raise.
    commercial_evaluation_status       → "Qualified", "Not Qualified", or "Query to be raised".
    reason_for_rejection               → Reason if not qualified, else "Not Applicable".

RULE: If a value is not found in the context, write "Not Found in Documents" — do NOT leave blank.
""",
    },
]

# ── Prompt template ──────────────────────────────────────────
PROMPT_TEMPLATE = PromptTemplate.from_template(
    """You are an expert Tender Evaluation AI Agent working for CPCL (Chennai Petroleum Corporation Limited).

Evaluation Section: {section_label}

TASK — Fill every single field listed below. For each field:
- If found in the documents, extract the exact value.
- If NOT found, write "Not Found in Documents".
- Never leave a field empty.

{instruction}

Bidder Documentation Context (retrieved from submitted documents):
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{context}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

CRITICAL: Your output MUST be in the exact structured JSON format. Fill ALL fields.
"""
)


# ── LLM factory ─────────────────────────────────────────────
def get_llm():
    """
    Priority:
    1. Vertex AI via API Key (using ChatVertexAI)
    2. Google AI Studio (using ChatGoogleGenerativeAI)
    3. Ollama local fallback
    """
    api_key = os.getenv("GEMINI_API_KEY", "")
    model_id = os.getenv("MODEL_ID", "gemini-3-flash-preview")
    gcp_project = os.getenv("GOOGLE_CLOUD_PROJECT", "")
    gcp_location = os.getenv("GOOGLE_CLOUD_LOCATION", "us-central1")
    if gcp_location == "global": gcp_location = "us-central1"
    
    # ── 1. Vertex AI via API Key ──────────────────────────────
    if gcp_project and api_key:
        try:
            from langchain_google_vertexai import ChatVertexAI
            logger.info(f"[LLM] Using {model_id} via Vertex AI + API Key (project={gcp_project})")
            # In latest langchain-google-vertexai, you can pass api_key directly
            llm = ChatVertexAI(
                model=model_id,
                api_key=api_key,
                project=gcp_project,
                location=gcp_location,
                temperature=0,
            )
            return llm, f"{model_id} (Vertex AI)"
        except Exception as exc:
            logger.warning(f"[LLM] Vertex AI + API Key failed: {exc}")

    # ── 2. Google AI Studio (direct API key) ──────────────────
    if api_key:
        try:
            from langchain_google_genai import ChatGoogleGenerativeAI
            logger.info(f"[LLM] Using {model_id} via Google AI Studio (API Key)")
            llm = ChatGoogleGenerativeAI(
                model=model_id,
                google_api_key=api_key,
                temperature=0,
            )
            return llm, f"{model_id} (Google AI Studio)"
        except Exception as exc:
            logger.warning(f"[LLM] Google AI Studio init failed: {exc}")

    # ── 3. Ollama local fallback ──────────────────────────────
    logger.info("[LLM] Using Ollama qwen2.5-coder (local fallback)")
    return ChatOllama(model="qwen2.5-coder:latest", temperature=0), "Ollama qwen2.5-coder"

    # ── 3. Ollama local fallback ──────────────────────────────
    logger.info("[LLM] Using Ollama qwen2.5-coder (local fallback)")
    return ChatOllama(model="qwen2.5-coder:latest", temperature=0), "Ollama qwen2.5-coder"


# ── Single section evaluator ─────────────────────────────────
def run_single_evaluation(section: dict, vectorstore: PGVector, llm, llm_label: str) -> dict:
    sid   = section["id"]
    label = section["label"]
    logger.info(f"▶  [{sid.upper()}] {label}")

    docs = vectorstore.similarity_search(section["search_query"], k=15)
    if not docs:
        logger.warning(f"   No documents retrieved for section {sid}.")
        return {}

    context_text = "\n\n".join(d.page_content for d in docs)
    logger.info(f"   Retrieved {len(docs)} chunks.")

    formatted = PROMPT_TEMPLATE.format(
        section_label=label,
        instruction=section["instruction"],
        context=context_text,
    )

    try:
        # Use with_structured_output for the primary LLM
        structured_llm = llm.with_structured_output(section["schema"])
        result = structured_llm.invoke(formatted)
    except Exception as exc:
        logger.warning(f"   Primary LLM failed: {exc}. Trying Ollama fallback...")
        fallback = ChatOllama(model="qwen2.5-coder:latest", temperature=0)
        structured_fallback = fallback.with_structured_output(section["schema"])
        try:
            result = structured_fallback.invoke(formatted)
        except Exception as fb_exc:
            logger.error(f"   Fallback also failed: {fb_exc}")
            return {}

    if result is None:
        logger.error(f"   Section {sid} — LLM returned None.")
        return {}

    data = result.model_dump()
    logger.info(f"   ✓ Section {sid.upper()} complete ({len(data)} fields extracted).")
    return data


# ── Excel builder ────────────────────────────────────────────
CRITERIA_FILES = {
    "a": "output-excel-criterias/criteriaa.md",
    "b": "output-excel-criterias/criteriab.md",
    "c": "output-excel-criterias/criteriac.md",
    "d": "output-excel-criterias/criteriad.md",
    "e": "output-excel-criterias/criteriae.md",
    "f": "output-excel-criterias/criteriaf.md",
}

# Colours
HDR_FILL   = PatternFill("solid", fgColor="1F3864")   # dark navy
HDR_FONT   = Font(color="FFFFFF", bold=True, size=11)
TITLE_FILL = PatternFill("solid", fgColor="D5E8F7")   # light blue
FIELD_ODD  = PatternFill("solid", fgColor="D6E4F0")
FIELD_EVN  = PatternFill("solid", fgColor="AED6F1")
VAL_ODD    = PatternFill("solid", fgColor="EBF5FB")
VAL_EVN    = PatternFill("solid", fgColor="FDFEFE")
THIN       = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


def _criteria_rows(section_id: str) -> list[dict]:
    """Parse the criteria MD and return list of {sl_no, item_description, tender_requirement} rows."""
    md_path = Path(CRITERIA_FILES.get(section_id, ""))
    if not md_path.exists():
        return []
    rows = []
    for line in md_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line.startswith("|"):
            continue
        cols = [c.strip() for c in line.strip("|").split("|")]
        if len(cols) >= 2 and cols[0] not in ("", "---", "Sl. No.", "Tender Requirement"):
            rows.append({
                "Sl. No.":            cols[0] if len(cols) > 0 else "",
                "Item Description":   cols[1] if len(cols) > 1 else "",
                "Tender Requirement": cols[2] if len(cols) > 2 else "",
            })
    return rows


def save_to_excel(all_results: dict, output_dir: Path, bidder_label: str):
    excel_path = output_dir / f"{bidder_label}_evaluation.xlsx"

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        for section in EVALUATIONS:
            sid   = section["id"]
            label = section["label"]
            data  = all_results.get(sid, {})

            # Build rows: merge criteria table with AI extracted values
            criteria_rows = _criteria_rows(sid)

            # Build a flat display: Sl.No | Item Description | Tender Requirement | Bidder Value
            display_rows = []
            for crow in criteria_rows:
                item = crow["Item Description"]
                # Try to match extracted field by checking if any key roughly maps to this item
                extracted_val = "—"
                if data:
                    # Attempt rough key lookup
                    item_key = re.sub(r"[^a-z0-9]", "_", item.lower())[:30]
                    for k, v in data.items():
                        if any(word in k for word in item_key.split("_") if len(word) > 3):
                            extracted_val = "; ".join(str(x) for x in v) if isinstance(v, list) else str(v)
                            break
                display_rows.append({
                    "Sl. No.":            crow["Sl. No."],
                    "Item Description":   item,
                    "Tender Requirement": crow["Tender Requirement"],
                    "Bidder Value":       extracted_val,
                })

            # Append a separator, then all extracted fields
            if data:
                display_rows.append({"Sl. No.": "", "Item Description": "─── AI EXTRACTED DATA ───", "Tender Requirement": "", "Bidder Value": ""})
                for k, v in data.items():
                    val = "; ".join(str(x) for x in v) if isinstance(v, list) else str(v)
                    display_rows.append({
                        "Sl. No.":            "",
                        "Item Description":   k.replace("_", " ").title(),
                        "Tender Requirement": "",
                        "Bidder Value":       val,
                    })

            df = pd.DataFrame(display_rows)
            # Sheet name max 31 chars
            sheet_name = f"({sid}) {label[4:26]}"
            df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)

            ws = writer.sheets[sheet_name]

            # Title rows
            ws["A1"] = label
            ws["A1"].font      = Font(bold=True, size=13, color="1F3864")
            ws["A1"].fill      = TITLE_FILL
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws.merge_cells("A1:D1")
            ws.row_dimensions[1].height = 26

            ws["A2"] = f"Bidder: {bidder_label}"
            ws["A2"].font      = Font(bold=True, size=11, color="1F3864")
            ws["A2"].alignment = Alignment(horizontal="left")
            ws.row_dimensions[2].height = 18

            # Header row (row 3 due to startrow=2)
            for cell in ws[3]:
                cell.fill      = HDR_FILL
                cell.font      = HDR_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border    = THIN
            ws.row_dimensions[3].height = 22

            # Data rows
            for row_idx, row in enumerate(ws.iter_rows(min_row=4, max_row=ws.max_row), start=0):
                is_odd = row_idx % 2 == 0
                for col_idx, cell in enumerate(row):
                    cell.border    = THIN
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    if col_idx == 0:
                        cell.fill = FIELD_ODD if is_odd else FIELD_EVN
                        cell.font = Font(bold=False, size=9)
                    elif col_idx == 1:
                        cell.fill = FIELD_ODD if is_odd else FIELD_EVN
                        cell.font = Font(bold=True, size=10)
                    elif col_idx == 2:
                        cell.fill = VAL_ODD if is_odd else VAL_EVN
                        cell.font = Font(italic=True, size=10, color="555555")
                    else:
                        cell.fill = VAL_ODD if is_odd else VAL_EVN
                        cell.font = Font(size=10, bold=False)

            # Column widths
            ws.column_dimensions["A"].width = 10
            ws.column_dimensions["B"].width = 45
            ws.column_dimensions["C"].width = 35
            ws.column_dimensions["D"].width = 55

    logger.info(f"  ✅ Excel saved → {excel_path}")
    return excel_path


# ── Orchestrator ─────────────────────────────────────────────
def evaluate_bidder(bidder_name: str):
    # Derive folder name: "Bidder2 offer.zip" → "Bidder2offerevaluation"
    clean      = re.sub(r"[^a-zA-Z0-9]", "", Path(bidder_name).stem)
    output_dir = Path(clean + "evaluation")
    output_dir.mkdir(exist_ok=True)
    logger.info(f"Output directory: {output_dir.resolve()}")

    embeddings  = OllamaEmbeddings(model="nomic-embed-text")
    vectorstore = PGVector(
        embeddings=embeddings,
        collection_name=COLLECTION_NAME,
        connection=CONNECTION_STRING,
    )

    llm, llm_label = get_llm()
    all_results     = {}

    for section in EVALUATIONS:
        sid  = section["id"]
        data = run_single_evaluation(section, vectorstore, llm, llm_label)

        json_path = output_dir / f"section_{sid}.json"
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=4, ensure_ascii=False)
        logger.info(f"   JSON → {json_path}")

        all_results[sid] = data

    save_to_excel(all_results, output_dir, clean)

    logger.info("\n" + "=" * 60)
    logger.info(f"ALL 6 EVALUATIONS COMPLETE — {bidder_name}")
    logger.info(f"Output folder : {output_dir.resolve()}")
    logger.info("=" * 60)


# ── Entry point ──────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Full 6-section bidder evaluation")
    parser.add_argument(
        "--bidder",
        type=str,
        default="Bidder2 offer.zip",
        help="Bidder zip file or folder name (used for output naming only)",
    )
    args = parser.parse_args()
    evaluate_bidder(args.bidder)
